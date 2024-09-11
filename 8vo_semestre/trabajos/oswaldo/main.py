import os
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
from copy import copy

# Función para generar las fechas de sesiones
def generar_fechas(mes, año, dias_laborables):
    fechas = []
    inicio = datetime(año, mes, 1)
    while inicio.month == mes:
        if inicio.weekday() in dias_laborables:
            fechas.append(inicio.strftime("%Y-%m-%d"))
        inicio += timedelta(days=1)
    return fechas

# Función para copiar el contenido y estilos de una hoja a otra
def copiar_hoja(origen, destino):
    # Copiar los valores y estilos de las celdas
    for fila in origen.iter_rows():
        for celda in fila:
            nueva_celda = destino.cell(row=celda.row, column=celda.column, value=celda.value)
            if celda.has_style:
                nueva_celda.font = copy(celda.font)
                nueva_celda.border = copy(celda.border)
                nueva_celda.fill = copy(celda.fill)
                nueva_celda.number_format = celda.number_format
                nueva_celda.protection = copy(celda.protection)
                nueva_celda.alignment = copy(celda.alignment)
    
    # Copiar celdas combinadas
    if origen.merged_cells:
        for range_ in origen.merged_cells.ranges:
            destino.merge_cells(range_.coord)
    
    # Copiar el ancho de las columnas
    for col in origen.column_dimensions:
        destino.column_dimensions[col].width = origen.column_dimensions[col].width
    
    # Copiar la altura de las filas
    for row in origen.row_dimensions:
        destino.row_dimensions[row].height = origen.row_dimensions[row].height

# Configuraciones de los meses y días laborables
mesociclos = [
    {"nombre": "mesociclo-1-febrero", "mes": 2, "año": 2024, "dias_laborables": [0, 1, 2], "carpeta": "meso1"},
    {"nombre": "mesociclo-2-marzo", "mes": 3, "año": 2024, "dias_laborables": [0, 1, 2], "carpeta": "meso2"},
    {"nombre": "mesociclo-3-abril", "mes": 4, "año": 2024, "dias_laborables": [0, 1, 2], "carpeta": "meso3"},
    {"nombre": "mesociclo-3-mayo", "mes": 5, "año": 2024, "dias_laborables": [0, 1, 2, 3, 4], "carpeta": "meso3", "excepciones": ["2024-05-13"]}
]

# Crear los archivos de Excel
for mesociclo in mesociclos:
    wb = Workbook()
    mes = mesociclo["mes"]
    año = mesociclo["año"]
    print(f"{mes}-{año}")
    dias_laborables = mesociclo["dias_laborables"]
    excepciones = mesociclo.get("excepciones", [])
    carpeta = mesociclo["carpeta"]
    
    fechas = generar_fechas(mes, año, dias_laborables)
    
    # Eliminar excepciones de las fechas generadas
    fechas = [fecha for fecha in fechas if fecha not in excepciones]
    
    # Obtener la lista de archivos en la carpeta correspondiente
    archivos_sesiones = sorted([f for f in os.listdir(carpeta) if f.endswith('.xlsx')])
    
    # Crear las hojas de cálculo y copiar el contenido de las sesiones
    for i, fecha in enumerate(fechas):
        print(fecha)
        ws = wb.create_sheet(title=fecha)
        archivo_origen = archivos_sesiones[min(i, len(archivos_sesiones) - 1)]
        wb_origen = load_workbook(os.path.join(carpeta, archivo_origen))
        ws_origen = wb_origen.active
        copiar_hoja(ws_origen, ws)
        
        # Añadir la fecha de la sesión en la celda D16
        ws["D16"] = fecha

        # Establecer el valor 29 en todas las celdas combinadas en columnas N a O y fila 16
        rango_combinado = "N16:O16"
        ws.merge_cells(rango_combinado)
        ws["N16"] = 29
        rango_combinado = "K39:O39"
        ws.merge_cells(rango_combinado)
        ws["K39"] = "MYCIM JHON EDINSON RIVERA TAPIAS"
        rango_combinado = "K40:O40"
        ws.merge_cells(rango_combinado)
        ws["K40"] = "JDIMYF"
    
    # Eliminar la hoja por defecto
    del wb["Sheet"]
    
    # Guardar el archivo de Excel
    nombre_archivo = f"{mesociclo['nombre']}.xlsx"
    wb.save(nombre_archivo)
    print(f"Archivo '{nombre_archivo}' creado con éxito.")
